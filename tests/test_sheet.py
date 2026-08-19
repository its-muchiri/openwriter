from openpyxl import Workbook, load_workbook
from pathlib import Path
from wp_ai_publisher.sheet import KeywordSheet
def test_pending_and_writeback():
    folder = Path(".test-artifacts/sheet"); folder.mkdir(parents=True, exist_ok=True); path = folder / "keywords.xlsx"; wb = Workbook(); ws = wb.active; ws.append(["keyword", "site_id", "status"]); ws.append(["tea", "a", None]); ws.append(["done", "a", "done"]); wb.save(path)
    sheet = KeywordSheet(path); rows = sheet.pending(); assert len(rows) == 1
    sheet.update(rows[0], status="done", post_id=7); assert load_workbook(path).active.cell(2, sheet.headers["post_id"]).value == 7
