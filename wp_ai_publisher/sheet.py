from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

REQUIRED = {"keyword", "site_id"}
MANAGED = ["status", "post_id", "edit_url", "last_run_at", "error_message"]

@dataclass
class SheetRow:
    index: int
    values: dict

class KeywordSheet:
    def __init__(self, path: Path):
        self.path = path; self.book = load_workbook(path); self.ws = self.book.active
        self.headers = {str(c.value).strip(): c.column for c in self.ws[1] if c.value}
        missing = REQUIRED - self.headers.keys()
        if missing: raise ValueError("Sheet missing required columns: " + ", ".join(sorted(missing)))
        for name in MANAGED:
            if name not in self.headers:
                col = self.ws.max_column + 1; self.ws.cell(1, col, name); self.headers[name] = col

    def pending(self, site_id: str | None = None, limit: int | None = None):
        found = []
        for i in range(2, self.ws.max_row + 1):
            values = {name: self.ws.cell(i, col).value for name, col in self.headers.items()}
            if not values["keyword"]: continue
            if values.get("status") not in (None, "", "pending", "error"): continue
            if site_id and values["site_id"] != site_id: continue
            found.append(SheetRow(i, values))
            if limit and len(found) >= limit: break
        return found

    def update(self, row: SheetRow, **fields):
        fields["last_run_at"] = datetime.now(timezone.utc).isoformat()
        for key, value in fields.items(): self.ws.cell(row.index, self.headers[key], value)
        self.book.save(self.path)
